from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q
from django.utils.text import slugify
import logging
from .models import Trainee, Task, SignOff, UnsignLog, Cohort

# Security logger for audit trail
security_logger = logging.getLogger('security')

@login_required
def trainee_list(request):
    """Display list of trainees in the current cohort"""
    # Handle view mode preference
    view_mode = request.GET.get('view', request.session.get('view_mode', 'standard'))
    request.session['view_mode'] = view_mode

    # Get current cohort
    current_cohort = Cohort.get_current_cohort()

    # Filter trainees by current cohort
    if current_cohort:
        trainees = Trainee.objects.filter(is_active=True, cohort=current_cohort).order_by('badge_number')
    else:
        trainees = Trainee.objects.filter(is_active=True).order_by('badge_number')

    # Get all active tasks for bulk operations
    tasks = Task.objects.filter(is_active=True).order_by('order')

    context = {
        'trainees': trainees,
        'view_mode': view_mode,
        'current_cohort': current_cohort,
        'tasks': tasks,
    }
    return render(request, 'tracker/trainee_list.html', context)


@login_required
def export_cohort_excel(request, cohort_id=None):
    """Export cohort data to Excel using the blank template"""
    from django.http import HttpResponse
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment
    from openpyxl.cell.cell import MergedCell
    import os

    # Get cohort
    if cohort_id:
        cohort = get_object_or_404(Cohort, id=cohort_id)
    else:
        cohort = Cohort.get_current_cohort()
        if not cohort:
            messages.error(request, 'No current cohort found')
            return redirect('trainee_list')

    # Load blank template
    template_path = 'Check list Orientation Blank.xlsx'
    if not os.path.exists(template_path):
        messages.error(request, 'Excel template not found')
        return redirect('trainee_list')

    wb = load_workbook(template_path)
    ws = wb.active

    # Update cohort name in header (T2 is top-left of merged range T2:U2)
    ws.cell(row=2, column=20, value=cohort.name)

    # Task order to Excel column mapping
    # Based on template structure
    task_column_map = {
        1: 3,   # Onboarding Process Brief -> Column C (3)
        2: 4,   # Police Clearance Form -> Column D (4)
        3: 7,   # Document Release Form/NDA -> Column G (7)
        4: 8,   # U.S. Citizen / ID Check -> Column H (8)
        5: 9,   # Read SOP's 208, 210 -> Column I (9)
        6: 10,  # Read SOP 501 A&B -> Column J (10)
        7: 11,  # Read SOP's 505, 506, 508, 509, 510 -> Column K (11)
        8: 12,  # Read SOP's 600, 601 -> Column L (12)
        9: 13,  # Onboarding Tour -> Column M (13)
        10: 14, # Onboarding Tour Quiz -> Column N (14)
        11: 15, # Read R.G. 8.13, R.G. 8.29 -> Column O (15)
        12: 16, # ALARA Statement -> Column P (16)
        13: 17, # Radiation Safety PowerPoint & Video -> Column Q (17)
        14: 18, # Reg Guides 8.13 / 8.29 Quiz -> Column R (18)
        15: 19, # Review Deficiencies -> Column S (19)
    }

    # Get all trainees for this cohort
    trainees = Trainee.objects.filter(
        cohort=cohort,
        is_active=True
    ).order_by('badge_number')

    # Get all tasks
    tasks = Task.objects.filter(is_active=True).order_by('order')

    # Detect sections in the template dynamically
    # Find all rows with "Badge Number" header (indicates section start)
    sections = []
    for row in range(1, 100):
        cell = ws.cell(row=row, column=1)
        if cell.value and 'Badge' in str(cell.value):
            # Data starts 3 rows after header
            data_start_row = row + 3

            # Count consecutive rows with badge number pattern
            data_row_count = 0
            for check_row in range(data_start_row, data_start_row + 30):
                check_cell = ws.cell(row=check_row, column=1)
                if check_cell.value and '#' in str(check_cell.value):
                    data_row_count += 1
                else:
                    break

            sections.append({
                'header_row': row,
                'data_start': data_start_row,
                'capacity': data_row_count
            })

            # Also update cohort name in this section's header
            ws.cell(row=row, column=20, value=cohort.name)

    # Build list of available row positions across all sections
    available_rows = []
    for section in sections:
        for i in range(section['capacity']):
            available_rows.append(section['data_start'] + i)

    # Write trainee data to available rows
    trainees_written = 0
    for idx, trainee in enumerate(trainees):
        if idx >= len(available_rows):
            # More trainees than template capacity
            messages.warning(
                request,
                f'Template capacity exceeded: {len(trainees)} trainees but only {len(available_rows)} rows available. '
                f'Only first {len(available_rows)} trainees exported.'
            )
            break

        current_row = available_rows[idx]

        # Column A: Badge number
        cell_a = ws.cell(row=current_row, column=1)
        # Skip if this row is in a merged cell area
        if isinstance(cell_a, MergedCell):
            continue
        cell_a.value = trainee.badge_number

        # Column B: Name (Last, First)
        cell_b = ws.cell(row=current_row, column=2)
        if not isinstance(cell_b, MergedCell):
            cell_b.value = trainee.full_name

        # Get all sign-offs for this trainee with staff information
        signoffs = SignOff.objects.filter(trainee=trainee).select_related('signed_by__staff_profile')
        signoff_dict = {signoff.task_id: signoff for signoff in signoffs}

        # Fill in sign-offs with staff initials
        for task in tasks:
            if task.id in signoff_dict:
                signoff = signoff_dict[task.id]

                # Get staff initials, fallback to 'X' if unavailable
                try:
                    initials = signoff.signed_by.staff_profile.initials
                except (AttributeError, Exception):
                    # Fallback to 'X' if staff profile is missing or other error
                    initials = 'X'

                # Write initials to corresponding column
                excel_col = task_column_map.get(task.order)
                if excel_col:
                    cell = ws.cell(row=current_row, column=excel_col)
                    # Only write if not a merged cell
                    if not isinstance(cell, MergedCell):
                        cell.value = initials
                        cell.alignment = Alignment(horizontal='center', vertical='center')

        trainees_written += 1

    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    # Sanitize filename to prevent path traversal attacks
    safe_cohort_name = slugify(cohort.name)
    filename = f'Check_list_Orientation_{safe_cohort_name}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    # Save workbook to response
    wb.save(response)

    return response


@login_required
def trainee_detail(request, badge_number):
    """Display detailed progress for a specific trainee"""
    trainee = get_object_or_404(Trainee, badge_number=badge_number)
    tasks = Task.objects.filter(is_active=True).prefetch_related('authorized_signers').order_by('order')
    signoffs = SignOff.objects.filter(trainee=trainee).select_related('task', 'signed_by')

    # Create a dict of task_id -> signoff for easier template lookup
    signoff_dict = {so.task.id: so for so in signoffs}

    # Combine tasks with their signoff status and permissions
    task_progress = []
    for task in tasks:
        signoff = signoff_dict.get(task.id)
        # Can unsign if: 1) has signoff, 2) is staff, AND 3) either is superuser OR authorized for this task
        can_unsign = (
            signoff is not None and
            (request.user.is_staff or request.user.is_superuser) and
            (request.user.is_superuser or task.can_user_sign_off(request.user))
        )
        task_progress.append({
            'task': task,
            'signoff': signoff,
            'can_sign_off': task.can_user_sign_off(request.user),
            'can_unsign': can_unsign,
        })

    context = {
        'trainee': trainee,
        'task_progress': task_progress,
        'progress_percentage': trainee.get_progress_percentage(),
    }
    return render(request, 'tracker/trainee_detail.html', context)

@login_required
def sign_off_task(request, badge_number, task_id):
    """Sign off a task for a trainee"""
    if request.method == 'POST':
        trainee = get_object_or_404(Trainee, badge_number=badge_number)
        task = get_object_or_404(Task, id=task_id)

        # Check if user has staff profile with sign-off permission
        try:
            if not request.user.staff_profile.can_sign_off:
                messages.error(request, 'Your account does not have sign-off permissions. Contact an administrator.')
                return redirect('trainee_detail', badge_number=badge_number)
        except AttributeError:
            messages.error(request, 'Staff profile not found. Contact an administrator to set up your account.')
            return redirect('trainee_detail', badge_number=badge_number)

        # Check if user is authorized to sign off this task
        if not task.can_user_sign_off(request.user):
            messages.error(request, f'You are not authorized to sign off "{task.name}".')
            return redirect('trainee_detail', badge_number=badge_number)

        score = request.POST.get('score', '').strip()
        notes = request.POST.get('notes', '')

        # Validate notes length
        if len(notes) > 10000:
            messages.error(request, 'Notes exceed maximum length of 10,000 characters.')
            return redirect('trainee_detail', badge_number=badge_number)

        # Validate score if task requires it
        if task.requires_score and task.minimum_score is not None:
            if not score:
                messages.error(request, f'Score is required for "{task.name}". Minimum passing score: {task.minimum_score}')
                return redirect('trainee_detail', badge_number=badge_number)

            try:
                score_value = float(score)
                if score_value < float(task.minimum_score):
                    messages.error(request, f'Score {score} does not meet the minimum requirement of {task.minimum_score} for "{task.name}".')
                    return redirect('trainee_detail', badge_number=badge_number)
            except ValueError:
                messages.error(request, f'Invalid score format. Please enter a numeric value.')
                return redirect('trainee_detail', badge_number=badge_number)

        # Create or update signoff
        signoff, created = SignOff.objects.get_or_create(
            trainee=trainee,
            task=task,
            defaults={
                'signed_by': request.user,
                'score': score,
                'notes': notes
            }
        )

        if not created:
            signoff.signed_by = request.user
            signoff.score = score
            signoff.notes = notes
            signoff.save()

        messages.success(request, f'Task "{task.name}" signed off successfully.')
        return redirect('trainee_detail', badge_number=badge_number)

    return redirect('trainee_detail', badge_number=badge_number)


@login_required
def unsign_task(request, badge_number, task_id):
    """Remove a sign-off from a trainee's task (with audit trail)"""
    if request.method == 'POST':
        trainee = get_object_or_404(Trainee, badge_number=badge_number)
        task = get_object_or_404(Task, id=task_id)

        # Check if user has permission to unsign (staff/superuser only)
        if not (request.user.is_staff or request.user.is_superuser):
            messages.error(request, 'You do not have permission to remove sign-offs.')
            return redirect('trainee_detail', badge_number=badge_number)

        # Check if user is authorized for this specific task (superuser can override)
        if not request.user.is_superuser and not task.can_user_sign_off(request.user):
            messages.error(request, f'You are not authorized to remove sign-offs for "{task.name}".')
            return redirect('trainee_detail', badge_number=badge_number)

        try:
            signoff = SignOff.objects.get(trainee=trainee, task=task)

            # Validate reason length
            reason = request.POST.get('reason', '')
            if len(reason) > 10000:
                messages.error(request, 'Reason exceeds maximum length of 10,000 characters.')
                return redirect('trainee_detail', badge_number=badge_number)

            # Create audit log before deleting
            UnsignLog.objects.create(
                trainee=trainee,
                task=task,
                original_signed_by=signoff.signed_by,
                original_signed_at=signoff.signed_at,
                original_score=signoff.score,
                original_notes=signoff.notes,
                unsigned_by=request.user,
                reason=reason
            )

            # Delete the sign-off
            signoff.delete()

            # Log security event
            security_logger.info(
                'Unsign action: trainee=%s, task=%s, original_signer=%s, unsigned_by=%s',
                trainee.badge_number, task.name, signoff.signed_by.username if signoff.signed_by else 'Unknown',
                request.user.username
            )

            messages.success(request, f'Sign-off for "{task.name}" has been removed.')

        except SignOff.DoesNotExist:
            messages.error(request, 'This task has not been signed off yet.')

        return redirect('trainee_detail', badge_number=badge_number)

    return redirect('trainee_detail', badge_number=badge_number)


@login_required
def archive_list(request):
    """Display list of archived cohorts with global search capability"""
    from django.db.models import Count, Q as Q_agg

    current_cohort = Cohort.get_current_cohort()
    search_query = request.GET.get('search', '').strip()

    # Limit search query length to prevent performance issues
    if len(search_query) > 100:
        search_query = search_query[:100]

    # If search query provided, search across ALL trainees
    search_results = []
    if search_query:
        # Get total number of active tasks (only once, not per trainee)
        total_tasks = Task.objects.filter(is_active=True).count()

        # Search by badge number OR name (case-insensitive)
        # Annotate with signoff count to avoid N+1 queries
        trainees = Trainee.objects.filter(
            Q(badge_number__icontains=search_query) |
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query),
            is_active=True
        ).select_related('cohort').annotate(
            signoff_count=Count('signoffs', distinct=True)
        ).order_by('cohort__year', 'cohort__semester', 'badge_number')

        # Add progress percentage to each trainee using annotated count
        for trainee in trainees:
            if total_tasks == 0:
                progress = 0
            else:
                progress = round((trainee.signoff_count / total_tasks) * 100, 1)

            search_results.append({
                'trainee': trainee,
                'progress': progress,
            })

    # Get all cohorts except the current one
    if current_cohort:
        cohorts = Cohort.objects.exclude(id=current_cohort.id)
    else:
        cohorts = Cohort.objects.all()

    # Add trainee count to each cohort
    cohorts_with_counts = []
    for cohort in cohorts:
        trainee_count = Trainee.objects.filter(cohort=cohort, is_active=True).count()
        cohorts_with_counts.append({
            'cohort': cohort,
            'trainee_count': trainee_count,
            'is_past': cohort.end_date < current_cohort.end_date if current_cohort else False,
        })

    # Sort by year and semester (most recent first)
    cohorts_with_counts.sort(key=lambda x: (x['cohort'].year, x['cohort'].semester == 'Fall'), reverse=True)

    context = {
        'cohorts': cohorts_with_counts,
        'current_cohort': current_cohort,
        'search_query': search_query,
        'search_results': search_results,
    }
    return render(request, 'tracker/archive_list.html', context)


@login_required
def archive_detail(request, cohort_id):
    """Display trainees for a specific archived cohort"""
    cohort = get_object_or_404(Cohort, id=cohort_id)
    current_cohort = Cohort.get_current_cohort()

    # Handle view mode preference
    view_mode = request.GET.get('view', request.session.get('view_mode', 'standard'))
    request.session['view_mode'] = view_mode

    # Get trainees for this cohort
    trainees = Trainee.objects.filter(is_active=True, cohort=cohort).order_by('badge_number')

    # Get all active tasks for bulk operations
    tasks = Task.objects.filter(is_active=True).order_by('order')

    context = {
        'trainees': trainees,
        'view_mode': view_mode,
        'cohort': cohort,
        'current_cohort': current_cohort,
        'is_archive': True,
        'tasks': tasks,
    }
    return render(request, 'tracker/trainee_list.html', context)


@login_required
def bulk_sign_off(request):
    """
    Bulk sign-off multiple trainees on one task, or one trainee on multiple tasks.

    Expected POST data (JSON):
    {
        "trainee_ids": [1, 2, 3],  # Or single trainee for multi-task mode
        "task_ids": [5],           # Or multiple tasks for single trainee mode
        "scores": {"5": "95"},     # Task ID -> score mapping (if required)
        "notes": "Completed together"
    }

    Returns JSON:
    {
        "success": true,
        "created": 3,
        "updated": 1,
        "skipped": [{"trainee": "#2523", "task": "Quiz", "reason": "Already signed off"}],
        "errors": []
    }
    """
    from django.http import JsonResponse
    from django.db import transaction
    import json

    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST request required'}, status=405)

    # Check if user has staff profile with sign-off permission
    try:
        if not request.user.staff_profile.can_sign_off:
            return JsonResponse({'success': False, 'error': 'Your account does not have sign-off permissions'}, status=403)
    except AttributeError:
        return JsonResponse({'success': False, 'error': 'Staff profile not found. Contact an administrator'}, status=403)

    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON'}, status=400)

    trainee_ids = data.get('trainee_ids', [])
    task_ids = data.get('task_ids', [])
    scores = data.get('scores', {})  # Dict of task_id -> score
    notes = data.get('notes', '')

    # Validation
    if not trainee_ids or not task_ids:
        return JsonResponse({'success': False, 'error': 'Must select at least one trainee and one task'}, status=400)

    # Limit bulk operations to prevent DoS attacks
    if len(trainee_ids) > 100:
        return JsonResponse({'success': False, 'error': 'Maximum 100 trainees allowed per bulk operation'}, status=400)
    if len(task_ids) > 100:
        return JsonResponse({'success': False, 'error': 'Maximum 100 tasks allowed per bulk operation'}, status=400)

    # Validate notes length
    if len(notes) > 10000:
        return JsonResponse({'success': False, 'error': 'Notes exceed maximum length of 10,000 characters'}, status=400)

    results = {
        'success': True,
        'created': 0,
        'updated': 0,
        'skipped': [],
        'errors': []
    }

    try:
        with transaction.atomic():
            # Fetch trainees and tasks
            trainees = Trainee.objects.filter(id__in=trainee_ids, is_active=True)
            tasks = Task.objects.filter(id__in=task_ids, is_active=True).prefetch_related('authorized_signers')

            if trainees.count() != len(trainee_ids):
                return JsonResponse({'success': False, 'error': 'One or more trainees not found'}, status=404)

            if tasks.count() != len(task_ids):
                return JsonResponse({'success': False, 'error': 'One or more tasks not found'}, status=404)

            # Process each combination
            for trainee in trainees:
                for task in tasks:
                    # Check authorization
                    if not task.can_user_sign_off(request.user):
                        results['skipped'].append({
                            'trainee': trainee.badge_number,
                            'task': task.name,
                            'reason': 'Not authorized to sign off this task'
                        })
                        continue

                    # Validate score if required
                    score = scores.get(str(task.id), '')
                    if task.requires_score and task.minimum_score is not None:
                        if not score:
                            results['errors'].append({
                                'trainee': trainee.badge_number,
                                'task': task.name,
                                'error': f'Score required (minimum: {task.minimum_score})'
                            })
                            continue

                        try:
                            score_value = float(score)
                            if score_value < float(task.minimum_score):
                                results['errors'].append({
                                    'trainee': trainee.badge_number,
                                    'task': task.name,
                                    'error': f'Score {score} below minimum {task.minimum_score}'
                                })
                                continue
                        except ValueError:
                            results['errors'].append({
                                'trainee': trainee.badge_number,
                                'task': task.name,
                                'error': 'Invalid score format'
                            })
                            continue

                    # Create or update sign-off
                    signoff, created = SignOff.objects.update_or_create(
                        trainee=trainee,
                        task=task,
                        defaults={
                            'signed_by': request.user,
                            'score': score,
                            'notes': notes
                        }
                    )

                    if created:
                        results['created'] += 1
                    else:
                        results['updated'] += 1

            # If there were errors, rollback transaction
            if results['errors']:
                transaction.set_rollback(True)
                results['success'] = False
                results['error'] = f"{len(results['errors'])} validation errors occurred. No changes made."
            else:
                # Log successful bulk operation
                security_logger.info(
                    'Bulk sign-off: %d trainees, %d tasks, %d created, %d updated by %s',
                    len(trainee_ids), len(task_ids), results['created'], results['updated'], request.user.username
                )

    except Exception as e:
        security_logger.error('Bulk sign-off failed: %s by %s', str(e), request.user.username)
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

    return JsonResponse(results)


# ============================================================================
# Advanced Training Views
# ============================================================================

@login_required
def advanced_staff_list(request):
    """List all active advanced training staff"""
    from .models import AdvancedStaff, AdvancedTrainingType

    # Get all active staff
    staff_list = AdvancedStaff.objects.filter(is_active=True).prefetch_related('trainings', 'trainings__training_type')

    # Get all training types for table headers
    training_types = AdvancedTrainingType.objects.filter(is_active=True).order_by('order')

    # Build progress data for each staff member
    staff_progress = []
    for staff in staff_list:
        # Get all trainings for this staff member
        trainings_dict = {}
        for training in staff.trainings.all():
            key = training.training_type.id
            if key not in trainings_dict:
                trainings_dict[key] = []
            trainings_dict[key].append(training)

        staff_progress.append({
            'staff': staff,
            'trainings_dict': trainings_dict
        })

    context = {
        'staff_progress': staff_progress,
        'training_types': training_types,
        'page_title': 'Advanced Training - Active Staff',
    }
    return render(request, 'tracker/advanced_staff_list.html', context)


@login_required
def advanced_staff_removed(request):
    """List all removed advanced training staff"""
    from .models import AdvancedStaff, AdvancedTrainingType

    # Get all removed staff
    staff_list = AdvancedStaff.objects.filter(is_active=False).prefetch_related('trainings', 'trainings__training_type')

    # Get all training types for table headers
    training_types = AdvancedTrainingType.objects.filter(is_active=True).order_by('order')

    # Build progress data for each staff member
    staff_progress = []
    for staff in staff_list:
        # Get all trainings for this staff member
        trainings_dict = {}
        for training in staff.trainings.all():
            key = training.training_type.id
            if key not in trainings_dict:
                trainings_dict[key] = []
            trainings_dict[key].append(training)

        staff_progress.append({
            'staff': staff,
            'trainings_dict': trainings_dict
        })

    context = {
        'staff_progress': staff_progress,
        'training_types': training_types,
        'page_title': 'Advanced Training - Removed Staff',
        'is_removed': True,
    }
    return render(request, 'tracker/advanced_staff_list.html', context)


@login_required
def advanced_staff_detail(request, badge_number):
    """Detail view for one staff member's advanced training"""
    from .models import AdvancedStaff, AdvancedTrainingType

    staff = get_object_or_404(AdvancedStaff, badge_number=badge_number)

    # Get all training types
    training_types = AdvancedTrainingType.objects.filter(is_active=True).order_by('order')

    # Get all trainings for this staff member, organized by type
    trainings_by_type = {}
    for training in staff.trainings.all():
        key = (training.training_type.id, training.custom_type)
        trainings_by_type[key] = training

    # Build training progress list
    training_progress = []
    for training_type in training_types:
        if training_type.allows_custom_type:
            # For "Other Training" types, show all existing custom trainings
            custom_trainings = staff.trainings.filter(training_type=training_type)
            for training in custom_trainings:
                training_progress.append({
                    'training_type': training_type,
                    'training': training,
                    'custom_type': training.custom_type,
                })
        else:
            # Regular training type
            key = (training_type.id, '')
            training = trainings_by_type.get(key)
            training_progress.append({
                'training_type': training_type,
                'training': training,
                'custom_type': '',
            })

    context = {
        'staff': staff,
        'training_progress': training_progress,
        'page_title': f'Advanced Training - {staff.get_full_name()}',
    }
    return render(request, 'tracker/advanced_staff_detail.html', context)


@login_required
def export_advanced_excel(request):
    """Export active advanced training staff to Excel"""
    return _export_advanced_excel_internal(request, is_active=True)


@login_required
def export_advanced_excel_removed(request):
    """Export removed advanced training staff to Excel"""
    return _export_advanced_excel_internal(request, is_active=False)


def _export_advanced_excel_internal(request, is_active=True):
    """Internal function to export advanced training data to Excel"""
    from django.http import HttpResponse
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment
    import os

    # Load template
    template_path = 'ADV_TrainingStatus_WIP.xlsx'
    if not os.path.exists(template_path):
        messages.error(request, 'Excel template not found')
        return redirect('advanced_staff_main')

    wb = load_workbook(template_path)

    # Select appropriate sheet
    sheet_name = 'ADV' if is_active else 'ADV_Removed'
    if sheet_name not in wb.sheetnames:
        messages.error(request, f'Sheet "{sheet_name}" not found in template')
        return redirect('advanced_staff_main')

    ws = wb[sheet_name]

    # Column mapping (1-indexed)
    COL_BADGE = 1
    COL_LAST = 2
    COL_FIRST = 3
    COL_ROLE = 4
    # KP Training: Date(5), Apprvd(6), Term(7)
    # Escort Training: Date(8), Apprvd(9), Term(10)
    # ExpSamp Training: Date(11), Apprvd(12), Term(13)
    # Other Training: Type(14), Date(15), Apprvd(16), Term(17)
    # Other Training 2: Type(18), Date(19), Apprvd(20), Term(21)

    # Get training types
    kp_training = AdvancedTrainingType.objects.get(name='KP Training')
    escort_training = AdvancedTrainingType.objects.get(name='Escort Training')
    expsamp_training = AdvancedTrainingType.objects.get(name='ExpSamp Training')
    other_training = AdvancedTrainingType.objects.get(name='Other Training')
    other_training_2 = AdvancedTrainingType.objects.get(name='Other Training 2')

    # Training type to column mapping
    training_column_map = {
        kp_training.id: (5, 6, 7, None),  # (date_col, apprvd_col, term_col, type_col)
        escort_training.id: (8, 9, 10, None),
        expsamp_training.id: (11, 12, 13, None),
        other_training.id: (15, 16, 17, 14),  # Type is in col 14
        other_training_2.id: (19, 20, 21, 18),  # Type is in col 18
    }

    # Clear existing data (rows 3 onward, keep headers in rows 1-2)
    # Delete rows from bottom to top to avoid index shifting
    max_row = ws.max_row
    if max_row > 2:
        ws.delete_rows(3, max_row - 2)

    # Get staff with prefetched training data
    staff_list = AdvancedStaff.objects.filter(is_active=is_active).prefetch_related('trainings').order_by('badge_number')

    # Write staff data
    current_row = 3  # Start writing at row 3
    for staff in staff_list:
        # Write staff info
        ws.cell(row=current_row, column=COL_BADGE, value=staff.badge_number)
        ws.cell(row=current_row, column=COL_LAST, value=staff.last_name)
        ws.cell(row=current_row, column=COL_FIRST, value=staff.first_name)
        ws.cell(row=current_row, column=COL_ROLE, value=staff.role)

        # Build training dictionary for quick lookup
        training_dict = {}
        for training in staff.trainings.all():
            # Group by training_type_id and custom_type
            key = (training.training_type_id, training.custom_type)
            if key not in training_dict:
                training_dict[key] = []
            training_dict[key].append(training)

        # Write training data
        for training_type_id, (date_col, apprvd_col, term_col, type_col) in training_column_map.items():
            # For "Other" types, we need to handle custom types
            if training_type_id == other_training.id:
                # Get first Other Training record
                matching = [t for t in staff.trainings.all() if t.training_type_id == training_type_id]
                if matching:
                    training = matching[0]
                    if training.completion_date:
                        ws.cell(row=current_row, column=date_col, value=training.completion_date.strftime('%m/%d/%Y'))
                    if training.approver_initials:
                        ws.cell(row=current_row, column=apprvd_col, value=training.approver_initials)
                    if training.termination_date:
                        ws.cell(row=current_row, column=term_col, value=training.termination_date.strftime('%m/%d/%Y'))
                    if training.custom_type and type_col:
                        ws.cell(row=current_row, column=type_col, value=training.custom_type)

            elif training_type_id == other_training_2.id:
                # Get second Other Training record
                matching = [t for t in staff.trainings.all() if t.training_type_id == other_training.id]
                if len(matching) > 1:
                    training = matching[1]
                    if training.completion_date:
                        ws.cell(row=current_row, column=date_col, value=training.completion_date.strftime('%m/%d/%Y'))
                    if training.approver_initials:
                        ws.cell(row=current_row, column=apprvd_col, value=training.approver_initials)
                    if training.termination_date:
                        ws.cell(row=current_row, column=term_col, value=training.termination_date.strftime('%m/%d/%Y'))
                    if training.custom_type and type_col:
                        ws.cell(row=current_row, column=type_col, value=training.custom_type)
            else:
                # Standard training types (KP, Escort, ExpSamp)
                matching = [t for t in staff.trainings.all() if t.training_type_id == training_type_id]
                if matching:
                    training = matching[0]
                    if training.completion_date:
                        ws.cell(row=current_row, column=date_col, value=training.completion_date.strftime('%m/%d/%Y'))
                    if training.approver_initials:
                        ws.cell(row=current_row, column=apprvd_col, value=training.approver_initials)
                    if training.termination_date:
                        ws.cell(row=current_row, column=term_col, value=training.termination_date.strftime('%m/%d/%Y'))

        current_row += 1

    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'ADV_TrainingStatus_{"Active" if is_active else "Removed"}_{datetime.now().strftime("%Y%m%d")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    # Save workbook to response
    wb.save(response)

    return response


@login_required
def advanced_staff_main(request):
    """
    Main full-featured advanced training page with inline editing, filtering, and search.
    Similar to trainee_list but for advanced training staff.
    """
    from .models import AdvancedStaff, AdvancedTrainingType
    from django.db.models import Prefetch

    # Get filter parameters
    role_filter = request.GET.get('role', '')
    status_filter = request.GET.get('status', 'active')  # active, removed, all

    # Base queryset
    if status_filter == 'removed':
        staff_qs = AdvancedStaff.objects.filter(is_active=False)
    elif status_filter == 'all':
        staff_qs = AdvancedStaff.objects.all()
    else:  # active
        staff_qs = AdvancedStaff.objects.filter(is_active=True)

    # Apply role filter if specified
    if role_filter:
        staff_qs = staff_qs.filter(role=role_filter)

    # Prefetch trainings for efficiency
    staff_list = staff_qs.prefetch_related(
        Prefetch('trainings', queryset=AdvancedTraining.objects.select_related('training_type'))
    ).order_by('badge_number')

    # Get all training types for table headers
    training_types = AdvancedTrainingType.objects.filter(is_active=True).order_by('order')

    # Build progress data for each staff member
    staff_progress = []
    for staff in staff_list:
        # Organize trainings by type ID
        trainings_by_type = {}
        for training in staff.trainings.all():
            type_id = training.training_type.id
            if type_id not in trainings_by_type:
                trainings_by_type[type_id] = []
            trainings_by_type[type_id].append(training)

        staff_progress.append({
            'staff': staff,
            'trainings_by_type': trainings_by_type
        })

    # Get all roles for filter dropdown
    role_choices = AdvancedStaff.ROLE_CHOICES

    context = {
        'staff_progress': staff_progress,
        'training_types': training_types,
        'role_choices': role_choices,
        'role_filter': role_filter,
        'status_filter': status_filter,
        'page_title': 'Advanced Training Management',
    }
    return render(request, 'tracker/advanced_staff_main.html', context)


@login_required
def update_advanced_training(request):
    """
    AJAX endpoint to add or update an advanced training record.

    POST data (JSON):
    {
        "staff_id": 123,
        "training_type_id": 1,
        "custom_type": "Package",  // optional
        "completion_date": "2025-01-15",
        "approver_initials": "TB",
        "termination_date": "2026-01-15",  // optional
        "notes": "Additional info"
    }
    """
    from django.http import JsonResponse
    from .models import AdvancedStaff, AdvancedTrainingType, AdvancedTraining
    from datetime import datetime
    import json

    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST request required'}, status=405)

    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON'}, status=400)

    # Validate required fields
    staff_id = data.get('staff_id')
    training_type_id = data.get('training_type_id')

    if not staff_id or not training_type_id:
        return JsonResponse({'success': False, 'error': 'staff_id and training_type_id are required'}, status=400)

    try:
        staff = AdvancedStaff.objects.get(id=staff_id)
        training_type = AdvancedTrainingType.objects.get(id=training_type_id)
    except (AdvancedStaff.DoesNotExist, AdvancedTrainingType.DoesNotExist) as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=404)

    # Parse dates
    completion_date = None
    if data.get('completion_date'):
        try:
            completion_date = datetime.strptime(data['completion_date'], '%Y-%m-%d').date()
        except ValueError:
            return JsonResponse({'success': False, 'error': 'Invalid completion_date format (use YYYY-MM-DD)'}, status=400)

    termination_date = None
    if data.get('termination_date'):
        try:
            termination_date = datetime.strptime(data['termination_date'], '%Y-%m-%d').date()
        except ValueError:
            return JsonResponse({'success': False, 'error': 'Invalid termination_date format (use YYYY-MM-DD)'}, status=400)

    # Validate termination date is after completion date
    if completion_date and termination_date and termination_date <= completion_date:
        return JsonResponse({'success': False, 'error': 'Termination date must be after completion date'}, status=400)

    custom_type = data.get('custom_type', '').strip()
    approver_initials = data.get('approver_initials', '').strip()
    notes = data.get('notes', '').strip()

    # Create or update training record
    training, created = AdvancedTraining.objects.update_or_create(
        staff=staff,
        training_type=training_type,
        custom_type=custom_type,
        defaults={
            'completion_date': completion_date,
            'approver_initials': approver_initials,
            'termination_date': termination_date,
            'notes': notes,
        }
    )

    return JsonResponse({
        'success': True,
        'created': created,
        'training': {
            'id': training.id,
            'completion_date': training.completion_date.isoformat() if training.completion_date else None,
            'approver_initials': training.approver_initials,
            'termination_date': training.termination_date.isoformat() if training.termination_date else None,
            'is_expired': training.is_expired,
            'is_expiring_soon': training.is_expiring_soon(),
        }
    })


@login_required
def delete_advanced_training(request, training_id):
    """AJAX endpoint to delete an advanced training record"""
    from django.http import JsonResponse
    from .models import AdvancedTraining

    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST request required'}, status=405)

    try:
        training = AdvancedTraining.objects.get(id=training_id)
        training.delete()
        return JsonResponse({'success': True})
    except AdvancedTraining.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Training record not found'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)
