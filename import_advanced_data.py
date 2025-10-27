#!/usr/bin/env python
"""
Import Advanced Training Data from Excel
==========================================
This script imports advanced training data from ADV_TrainingStatus_WIP.xlsx
into the Django database.

Usage:
    python import_advanced_data.py

Requirements:
    - ADV_TrainingStatus_WIP.xlsx must be in the project root
    - Database migrations must be run first
    - Training types must be populated (migration 0017)
"""

import os
import sys
import django
from datetime import datetime

# Setup Django
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'trainee_tracker.settings')
django.setup()

from tracker.models import AdvancedStaff, AdvancedTrainingType, AdvancedTraining
from openpyxl import load_workbook


def parse_date(value):
    """Convert Excel date value to Python date object"""
    if value is None or value == '':
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, str):
        # Handle string dates like "~9/1/2023"
        value = value.replace('~', '').strip()
        try:
            return datetime.strptime(value, '%m/%d/%Y').date()
        except ValueError:
            try:
                return datetime.strptime(value, '%Y-%m-%d').date()
            except ValueError:
                print(f"Warning: Could not parse date: {value}")
                return None
    return None


def import_staff_from_sheet(ws, is_active=True):
    """
    Import staff from a worksheet.

    Args:
        ws: openpyxl worksheet
        is_active: True for ADV sheet, False for ADV_Removed sheet
    """
    staff_created = 0
    staff_updated = 0
    training_created = 0
    training_updated = 0

    # Get training types
    kp_training = AdvancedTrainingType.objects.get(name='KP Training')
    escort_training = AdvancedTrainingType.objects.get(name='Escort Training')
    expsamp_training = AdvancedTrainingType.objects.get(name='ExpSamp Training')
    other_training = AdvancedTrainingType.objects.get(name='Other Training')
    other_training_2 = AdvancedTrainingType.objects.get(name='Other Training 2')

    # Column mapping (1-indexed)
    COL_BADGE = 1
    COL_LAST = 2
    COL_FIRST = 3
    COL_ROLE = 4
    # KP Training: Date(5), Apprvd(6), Term(7)
    # Escort Training: Date(8), Apprvd(9), Term(10)
    # ExpSamp Training: Date(11), Apprvd(12), Term(13)
    # Other Training: Type(14), Date(15), Apprvd(16), Term(17)
    # Other Training 2: Type(18), Date(19), Apprvd(20), Term(21)

    # Start from row 3 (skip headers in rows 1-2)
    for row_num in range(3, ws.max_row + 1):
        badge = ws.cell(row=row_num, column=COL_BADGE).value

        # Skip empty rows
        if not badge:
            continue

        # Ensure badge is string
        badge = str(badge).strip()
        if not badge:
            continue

        # Get name and role
        last_name = str(ws.cell(row=row_num, column=COL_LAST).value or '').strip()
        first_name = str(ws.cell(row=row_num, column=COL_FIRST).value or '').strip()
        role = str(ws.cell(row=row_num, column=COL_ROLE).value or 'Other').strip()

        if not last_name or not first_name:
            print(f"Warning: Skipping badge {badge} - missing name")
            continue

        # Create or update staff member
        staff, created = AdvancedStaff.objects.update_or_create(
            badge_number=badge,
            defaults={
                'first_name': first_name,
                'last_name': last_name,
                'role': role if role in dict(AdvancedStaff.ROLE_CHOICES) else 'Other',
                'is_active': is_active
            }
        )

        if created:
            staff_created += 1
            print(f"Created staff: {badge} - {last_name}, {first_name}")
        else:
            staff_updated += 1
            print(f"Updated staff: {badge} - {last_name}, {first_name}")

        # Import training records
        training_data = [
            # (training_type, date_col, apprvd_col, term_col, type_col, custom_type_value)
            (kp_training, 5, 6, 7, None, ''),
            (escort_training, 8, 9, 10, None, ''),
            (expsamp_training, 11, 12, 13, None, ''),
            (other_training, 15, 16, 17, 14, ''),  # Type is in col 14
            (other_training_2, 19, 20, 21, 18, ''),  # Type is in col 18
        ]

        for training_type, date_col, apprvd_col, term_col, type_col, custom_type_default in training_data:
            # Get values from Excel
            completion_date = parse_date(ws.cell(row=row_num, column=date_col).value)
            approver = str(ws.cell(row=row_num, column=apprvd_col).value or '').strip()
            term_date = parse_date(ws.cell(row=row_num, column=term_col).value)

            # Get custom type if applicable
            custom_type = ''
            if type_col:
                custom_type = str(ws.cell(row=row_num, column=type_col).value or '').strip()

            # Only create training record if there's a completion date or approver
            if completion_date or approver:
                training, tr_created = AdvancedTraining.objects.update_or_create(
                    staff=staff,
                    training_type=training_type,
                    custom_type=custom_type,
                    defaults={
                        'completion_date': completion_date,
                        'approver_initials': approver,
                        'termination_date': term_date,
                    }
                )

                if tr_created:
                    training_created += 1
                else:
                    training_updated += 1

    return staff_created, staff_updated, training_created, training_updated


def main():
    """Main import function"""
    print("="*60)
    print("Advanced Training Data Import")
    print("="*60)
    print()

    # Check if Excel file exists
    excel_file = 'ADV_TrainingStatus_WIP.xlsx'
    if not os.path.exists(excel_file):
        print(f"ERROR: {excel_file} not found in current directory")
        print(f"Current directory: {os.getcwd()}")
        return

    # Load workbook
    print(f"Loading {excel_file}...")
    wb = load_workbook(excel_file, data_only=True)

    # Import from ADV sheet (active staff)
    print("\n" + "="*60)
    print("Importing from 'ADV' sheet (Active Staff)")
    print("="*60)
    if 'ADV' not in wb.sheetnames:
        print("ERROR: 'ADV' sheet not found in workbook")
        return

    ws_adv = wb['ADV']
    s_created, s_updated, t_created, t_updated = import_staff_from_sheet(ws_adv, is_active=True)

    print("\nADV Sheet Summary:")
    print(f"  Staff Created: {s_created}")
    print(f"  Staff Updated: {s_updated}")
    print(f"  Training Records Created: {t_created}")
    print(f"  Training Records Updated: {t_updated}")

    # Import from ADV_Removed sheet (removed staff)
    print("\n" + "="*60)
    print("Importing from 'ADV_Removed' sheet (Removed Staff)")
    print("="*60)
    if 'ADV_Removed' not in wb.sheetnames:
        print("WARNING: 'ADV_Removed' sheet not found - skipping")
    else:
        ws_removed = wb['ADV_Removed']
        s_created_r, s_updated_r, t_created_r, t_updated_r = import_staff_from_sheet(ws_removed, is_active=False)

        print("\nADV_Removed Sheet Summary:")
        print(f"  Staff Created: {s_created_r}")
        print(f"  Staff Updated: {s_updated_r}")
        print(f"  Training Records Created: {t_created_r}")
        print(f"  Training Records Updated: {t_updated_r}")

        # Add to totals
        s_created += s_created_r
        s_updated += s_updated_r
        t_created += t_created_r
        t_updated += t_updated_r

    # Final summary
    print("\n" + "="*60)
    print("IMPORT COMPLETE")
    print("="*60)
    print(f"Total Staff Created: {s_created}")
    print(f"Total Staff Updated: {s_updated}")
    print(f"Total Training Records Created: {t_created}")
    print(f"Total Training Records Updated: {t_updated}")
    print()
    print("Next steps:")
    print("  1. Verify data in admin: /admin/tracker/advancedstaff/")
    print("  2. Check training records: /admin/tracker/advancedtraining/")
    print("  3. Access advanced training views (when implemented)")
    print()


if __name__ == '__main__':
    try:
        main()
    except Exception as e:
        print(f"\nERROR: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
